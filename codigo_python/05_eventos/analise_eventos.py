#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análise por EVENTO — lê todas as planilhas auditáveis e produz
assets/data/eventos_analise.json com:

1) Catálogo de eventos canônicos (identificados pela DATA do pico, pois a
   numeração de eventos muda entre planilhas).
2) Por evento: dificuldade objetiva (NSE da persistência), desempenho das
   redes quando o evento foi TESTE, VALIDAÇÃO e no GERAL (mediana de NSE).
3) Séries completas (observado × RNA) de TODOS os eventos para os modelos
   com planilha auditável — hidrograma + dispersão por evento no site.
4) "Curiosidades do geral": leituras automáticas (evento mais difícil/fácil,
   onde as redes mais ganham da persistência, alerta de teste fácil).

Suporta planilhas no formato clássico (aba DADOS) e no formato VAR das
rotações 2h STZ (aba VAR: OUT2H NIV / RNA FINAL).

Uso:  python codigo_python/05_eventos/analise_eventos.py   (na raiz do repo)
"""
import datetime as _dt
import glob
import json
import os
from statistics import median

import openpyxl


WB_DIR = "assets/audit_workbooks"
OUT = "assets/data/eventos_analise.json"


def nse(pairs):
    obs = [o for o, _ in pairs]
    if len(obs) < 3:
        return None
    mo = sum(obs) / len(obs)
    den = sum((o - mo) ** 2 for o in obs)
    if den == 0:
        return None
    num = sum((o - p) ** 2 for o, p in pairs)
    return 1 - num / den


def _fl(v):
    try:
        return float(v)
    except Exception:
        return None


def _hdr_map(hdr):
    return {str(h or "").strip(): i for i, h in enumerate(hdr)}


def _pick(ix, *names):
    for nm in names:
        if nm in ix:
            return ix[nm]
    return None


def _normalize_conj(cj):
    s = str(cj or "").strip().lower()
    if s.startswith("trein"):
        return "Treino"
    if s.startswith("test"):
        return "Teste"
    return "Validacao"


def _finalize_events(evs):
    out = {}
    for ev, e in evs.items():
        rs = e["rows"]
        if not rs:
            continue
        peak = max(rs, key=lambda x: x[2])
        prna = [(o, p) for _, _, o, p, _ in rs if p is not None]
        pper = [(o, p) for _, _, o, _, p in rs if p is not None]
        pico_data = peak[0][:10]
        out[ev] = {
            "conj": e["conj"],
            "n": len(rs),
            "pico_cm": round(peak[2]),
            "pico_data": pico_data,
            "nse_rna": nse(prna),
            "nse_pers": nse(pper),
            "mae": round(sum(abs(o - p) for o, p in prna) / len(prna), 1) if prna else None,
            "rows": rs,
        }
    return out


def processa_dados(ws):
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").strip() for h in next(it)]
    ix = _hdr_map(hdr)
    i_ev = _pick(ix, "EVENTO")
    i_cj = _pick(ix, "CONJUNTO", "CONJUNTO_AUDITORIA")
    i_dh = _pick(ix, "DATA_HORA")
    i_na = _pick(ix, "NIVEL_ATUAL_CM")
    i_obs = _pick(ix, "OBSERVADO_CM_AUDITORIA", "OBSERVADO_CM")
    i_rna = _pick(ix, "RNA_CM")
    i_pers = _pick(ix, "PERS_BASE_NIVEL_ATUAL_CM", "PERS_BASE_CM", "NIVEL_ATUAL_CM")
    if None in (i_ev, i_cj, i_obs, i_rna):
        raise ValueError("aba DADOS sem colunas obrigatórias")
    evs = {}
    for r in it:
        try:
            ev = r[i_ev]
            obs = r[i_obs]
            if ev is None or obs is None:
                continue
            obs = float(obs)
            cj = str(r[i_cj] or "")
            if i_dh is not None and r[i_dh] is not None:
                dh = str(r[i_dh])
            else:
                # fallback se não houver DATA_HORA
                dh = ""
            na = _fl(r[i_na]) if i_na is not None else 0.0
            rna = _fl(r[i_rna])
            pers = _fl(r[i_pers]) if i_pers is not None else na
            e = evs.setdefault(ev, {"conj": cj, "rows": []})
            e["rows"].append((dh, na or 0.0, obs, rna, pers))
        except Exception:
            continue
    return _finalize_events(evs)


def processa_var(ws):
    """Formato das auditáveis 2h STZ (aba VAR)."""
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").strip() for h in next(it)]
    ix = _hdr_map(hdr)
    i_ev = _pick(ix, "EVENTO")
    i_cj = _pick(ix, "CONJUNTO")
    i_obs = _pick(ix, "OUT2H NIV", "OUT4H NIV", "OUT8H NIV", "OUT12H NIV")
    i_dif = _pick(ix, "OUT2H DIF", "OUT4H DIF", "OUT8H DIF", "OUT12H DIF")
    i_rna = _pick(ix, "RNA FINAL", "RNA_CM")
    i_ano = _pick(ix, "ANO")
    i_mes = _pick(ix, "MÊS", "MES")
    i_dia = _pick(ix, "DIA")
    i_hora = _pick(ix, "HORA")
    i_min = _pick(ix, "MINUTO")
    if None in (i_ev, i_cj, i_obs, i_rna, i_ano, i_mes, i_dia, i_hora):
        raise ValueError("aba VAR sem colunas obrigatórias")
    evs = {}
    for r in it:
        try:
            ev = r[i_ev]
            obs = r[i_obs]
            if ev is None or obs is None:
                continue
            obs = float(obs)
            cj = str(r[i_cj] or "")
            ano = int(float(r[i_ano]))
            mes = int(float(r[i_mes]))
            dia = int(float(r[i_dia]))
            hora = int(float(r[i_hora]))
            minuto = int(float(r[i_min] or 0)) if i_min is not None else 0
            dh = f"{ano:04d}-{mes:02d}-{dia:02d} {hora:02d}:{minuto:02d}"
            rna = _fl(r[i_rna])
            dif = _fl(r[i_dif]) if i_dif is not None else None
            # persistência de nível = nível atual (obs - Δ horizonte)
            na = (obs - dif) if dif is not None else obs
            pers = na
            e = evs.setdefault(ev, {"conj": cj, "rows": []})
            e["rows"].append((dh, na, obs, rna, pers))
        except Exception:
            continue
    return _finalize_events(evs)


def processa(fp):
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    try:
        if "DADOS" in wb.sheetnames:
            return processa_dados(wb["DADOS"])
        if "VAR" in wb.sheetnames:
            return processa_var(wb["VAR"])
        raise ValueError(f"sem aba DADOS/VAR ({wb.sheetnames})")
    finally:
        wb.close()


def model_key_from_path(fp):
    name = os.path.basename(fp)
    for suf in (
        "_AUDITAVEL_INPUTS_RNA.xlsx",
        "_AUDITAVEL.xlsx",
        ".xlsx",
    ):
        if name.endswith(suf):
            name = name[: -len(suf)]
            break
    # 2H_ALT__001_alt_STZ_... -> 001_alt_STZ_...
    if name.startswith(("2H_ALT__", "2H_CONV__", "4H_ALT__", "4H_CONV__", "8H_ALT__", "8H_CONV__", "12H_ALT__", "12H_CONV__")):
        name = name.split("__", 1)[1]
    # remove hash suffix __deadbeef
    if "__" in name and len(name.rsplit("__", 1)[-1]) in (8, 10):
        left, right = name.rsplit("__", 1)
        if all(c in "0123456789abcdef" for c in right.lower()):
            name = left
    return name


def main():
    # Santa Tereza: ignora planilhas de Muçum (prefixo MUC_), que têm análise própria.
    files = sorted(
        fp
        for fp in glob.glob(os.path.join(WB_DIR, "*.xlsx"))
        if "MUC_" not in os.path.basename(fp).upper()
    )
    print(f"{len(files)} planilhas (Santa Tereza; Muçum excluído)")
    canon = {}
    campeao_series = {}
    ok = err = 0
    for i, fp in enumerate(files):
        name = model_key_from_path(fp)
        try:
            evs = processa(fp)
            ok += 1
        except Exception as ex:
            err += 1
            print("erro", name, ex)
            continue
        for _ev, d in evs.items():
            pd = _dt.date.fromisoformat(d["pico_data"])
            key = None
            for k in canon:
                if abs((_dt.date.fromisoformat(k) - pd).days) <= 5:
                    key = k
                    break
            if key is None:
                key = d["pico_data"]
            c = canon.setdefault(
                key,
                {
                    "pico_cm": d["pico_cm"],
                    "n_modelos": 0,
                    "nse_por_conj": {"Treino": [], "Validacao": [], "Teste": []},
                    "nse_pers": [],
                },
            )
            c["pico_cm"] = max(c["pico_cm"], d["pico_cm"])
            c["n_modelos"] += 1
            cj = _normalize_conj(d["conj"])
            if d["nse_rna"] is not None:
                c["nse_por_conj"][cj].append(round(d["nse_rna"], 4))
            if d["nse_pers"] is not None:
                c["nse_pers"].append(round(d["nse_pers"], 4))
            if key not in campeao_series.get(name, {}):
                campeao_series.setdefault(name, {})[key] = {
                    "conj": cj,
                    "serie": [
                        [r[0], round(r[2]), round(r[3], 1) if r[3] is not None else None]
                        for r in d["rows"]
                    ],
                }
        if (i + 1) % 20 == 0:
            print(f"  {i + 1}/{len(files)}")

    eventos = []
    for key, c in sorted(canon.items()):
        med = lambda a: round(median(a), 3) if a else None
        eventos.append(
            {
                "pico_data": key,
                "pico_cm": c["pico_cm"],
                "n_modelos": c["n_modelos"],
                "dificuldade_nse_pers": med(c["nse_pers"]),
                "nse_teste": med(c["nse_por_conj"]["Teste"]),
                "nse_validacao": med(c["nse_por_conj"]["Validacao"]),
                "nse_treino": med(c["nse_por_conj"]["Treino"]),
                "n_teste": len(c["nse_por_conj"]["Teste"]),
                "n_validacao": len(c["nse_por_conj"]["Validacao"]),
            }
        )

    com_dif = [e for e in eventos if e["dificuldade_nse_pers"] is not None]
    cur = []
    if com_dif:
        dificil = min(com_dif, key=lambda e: e["dificuldade_nse_pers"])
        facil = max(com_dif, key=lambda e: e["dificuldade_nse_pers"])
        cur.append(
            f"O evento mais difícil de prever é o de {dificil['pico_data']} (pico {dificil['pico_cm']/100:.1f} m): a persistência só alcança NSE {dificil['dificuldade_nse_pers']:.2f}."
        )
        cur.append(
            f"O evento mais fácil é o de {facil['pico_data']}: a persistência sozinha chega a NSE {facil['dificuldade_nse_pers']:.2f} — bons resultados de teste nele merecem leitura cautelosa."
        )
        ganho = [
            (e, (e["nse_teste"] or 0) - (e["dificuldade_nse_pers"] or 0))
            for e in com_dif
            if e["nse_teste"] is not None
        ]
        if ganho:
            g = max(ganho, key=lambda x: x[1])
            cur.append(
                f"Onde as redes mais ganham da persistência: evento de {g[0]['pico_data']} (+{g[1]:.2f} de NSE sobre o baseline no teste)."
            )
        alerta = [e for e in com_dif if e["nse_teste"] is not None and e["dificuldade_nse_pers"] >= 0.9]
        if alerta:
            l = ", ".join(e["pico_data"] for e in alerta)
            cur.append(
                f"Atenção: quando o teste caiu em evento 'fácil' ({l}), o NSE alto pode refletir a facilidade do evento, não a força da rede — compare sempre com o NSE da persistência."
            )

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(
            {"gerado_de": ok, "eventos": eventos, "curiosidades": cur, "campeoes": campeao_series},
            f,
            ensure_ascii=False,
        )
    print(f"ok={ok} err={err} | eventos canônicos: {len(eventos)} | campeões com séries: {len(campeao_series)}")
    print("OUT:", OUT, os.path.getsize(OUT) // 1024, "KB")


if __name__ == "__main__":
    main()
