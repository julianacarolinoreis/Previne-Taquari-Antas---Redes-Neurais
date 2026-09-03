"""Núcleo do experimento RNA multi-horizonte (MIMO) — Santa Tereza.

Carrega dados auditáveis dos .mat principais, alinha amostras entre horizontes,
treina MLP multi-saída estilo PREVINE (logsig) e calcula métricas NASH/PERS/E95.
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import h5py
import numpy as np
from scipy.io import loadmat

ROOT = Path(__file__).resolve().parents[2]

MODELS = {
    "2h": {
        "mat": ROOT / "previne/assets/mat/009_alt_STZ_2H_R09_T10-15-16_V1-5-12-17-21.mat",
        "workbook": ROOT / "assets/audit_workbooks/2H_ALT__009_alt_STZ_2H_R09_T10-15-16_V1-5-12-17-21.xlsx",
        "sheet": "VAR",
        "horizon_h": 2,
        "n_inputs": 15,
    },
    "4h": {
        "mat": ROOT / "assets/mat/4H_ALT__V01_R00_BASELINE_nh52_nit10_cic100000.mat",
        "workbook": ROOT / "assets/audit_workbooks/4H_ALT__V01_R00_BASELINE_nh52_nit10_cic100000.xlsx",
        "sheet": "DADOS",
        "horizon_h": 4,
        "n_inputs": 26,
    },
    "8h": {
        "mat": ROOT / "previne/assets/mat/RNAPREV__SANTA_TEREZA__08h__ALT__V001__31inputs_63hiddens_20260821.mat",
        "horizon_h": 8,
        "n_inputs": 31,
        "delta_target": True,
    },
}


def logsig(z: np.ndarray) -> np.ndarray:
    z = np.clip(z, -60.0, 60.0)
    return 1.0 / (1.0 + np.exp(-z))


def _read_mat_dict(path: Path) -> dict:
    try:
        return loadmat(path, squeeze_me=True)
    except NotImplementedError:
        out: dict = {}
        with h5py.File(path, "r") as f:
            for key in f.keys():
                out[key] = np.array(f[key])
        return out


@dataclass
class HorizonDataset:
    name: str
    horizon_h: int
    inputs: np.ndarray
    atual: np.ndarray
    target_abs: np.ndarray
    delta: np.ndarray
    split: np.ndarray
    mat_path: Path
    events: np.ndarray | None = None

    @property
    def n_samples(self) -> int:
        return int(self.inputs.shape[0])

    @property
    def n_inputs(self) -> int:
        return int(self.inputs.shape[1])


def load_event_ids(name: str, n_samples: int) -> np.ndarray | None:
    cfg = MODELS[name]
    workbook = cfg.get("workbook")
    if not workbook or not Path(workbook).exists():
        return None
    import openpyxl

    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    sheet_name = cfg.get("sheet")
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h or "").strip() for h in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    if "EVENTO" not in ix:
        return None
    events = []
    for row in rows[1:]:
        if not row or row[ix["EVENTO"]] is None:
            continue
        events.append(int(float(row[ix["EVENTO"]])))
    if len(events) != n_samples:
        return None
    return np.asarray(events, dtype=int)


def load_horizon_dataset(name: str) -> HorizonDataset:
    cfg = MODELS[name]
    path = Path(cfg["mat"])
    m = _read_mat_dict(path)
    dados = np.asarray(m["DADOS"], float)
    if dados.ndim == 2 and dados.shape[0] < dados.shape[1] and dados.shape[0] <= 40:
        dados = dados.T
    n_in = int(cfg["n_inputs"])
    inputs = np.asarray(dados[:, :n_in], float)
    split = np.asarray(m["X"], float).ravel().astype(int)
    if cfg.get("delta_target"):
        delta = np.asarray(m["Tctot"], float).ravel()
        atual = inputs[:, 0].copy()
        target_abs = atual + delta
    else:
        target_abs = np.asarray(m["Tctot1"], float).ravel()
        atual = np.asarray(m["ATUAL_TOT"], float).ravel()
        delta = target_abs - atual
    if inputs.shape[0] != split.size:
        raise ValueError(f"{name}: split size {split.size} != samples {inputs.shape[0]}")
    events = load_event_ids(name, inputs.shape[0])
    return HorizonDataset(
        name=name,
        horizon_h=int(cfg["horizon_h"]),
        inputs=inputs,
        atual=atual,
        target_abs=target_abs,
        delta=delta,
        split=split,
        mat_path=path,
        events=events,
    )


def alignment_key(dataset: HorizonDataset) -> np.ndarray:
    return np.round(
        np.column_stack([dataset.atual, dataset.inputs[:, 0], dataset.inputs[:, 1], dataset.inputs[:, 2]]),
        1,
    )


def align_horizons(names: Iterable[str]) -> dict:
    datasets = [load_horizon_dataset(n) for n in names]
    index_maps = []
    for ds in datasets:
        mapping: dict[tuple, list[int]] = {}
        keys = alignment_key(ds)
        for idx, row in enumerate(keys):
            mapping.setdefault(tuple(row), []).append(idx)
        index_maps.append(mapping)
    common_keys = set(index_maps[0].keys())
    for mp in index_maps[1:]:
        common_keys &= set(mp.keys())
    rows = []
    for key in sorted(common_keys):
        indices = tuple(maps[key][0] for maps in index_maps)
        split = datasets[0].split[indices[0]]
        event = None
        if datasets[0].events is not None:
            event = int(datasets[0].events[indices[0]])
            # prefer agreement with 4h event id when available
            if len(datasets) > 1 and datasets[1].events is not None:
                event = int(datasets[1].events[indices[1]])
        rows.append({"key": key, "indices": indices, "split": int(split), "event": event})
    return {"datasets": datasets, "rows": rows}


def _scalar(v) -> float:
    return float(np.asarray(v).ravel()[0])


def load_mat_weights(path: Path) -> dict[str, np.ndarray]:
    m = _read_mat_dict(path)
    wh = np.atleast_2d(np.asarray(m["wh"], float))
    ws = np.asarray(m["ws"], float)
    return {
        "wh": wh,
        "bh": np.asarray(m["bh"], float).ravel(),
        "ws": ws,
        "bs": _scalar(m["bs"]),
        "ae": np.asarray(m["ae"], float).ravel(),
        "be": np.asarray(m["be"], float).ravel(),
        "au": _scalar(m["au"]),
        "bu": _scalar(m["bu"]),
    }


def predict_direct(weights: dict[str, np.ndarray], x: np.ndarray, atual: float) -> float:
    pn = (x - weights["be"]) / weights["ae"]
    h = logsig(pn @ weights["wh"].T + weights["bh"])
    yn = logsig(h @ weights["ws"] + weights["bs"])
    return float(atual + yn * weights["au"] + weights["bu"])


def predict_direct_batch(weights: dict[str, np.ndarray], x: np.ndarray, atual: np.ndarray) -> np.ndarray:
    pn = (x - weights["be"]) / weights["ae"]
    wh = weights["wh"]
    if wh.shape[1] == pn.shape[1]:
        h = logsig(pn @ wh.T + weights["bh"])
    elif wh.shape[0] == pn.shape[1]:
        h = logsig(pn @ wh + weights["bh"])
    else:
        raise ValueError(f"wh shape incompatível: {wh.shape} vs inputs {pn.shape[1]}")
    ws = weights["ws"]
    if ws.ndim == 1:
        yn = logsig(h @ ws + weights["bs"])
    else:
        yn = logsig(h @ ws.reshape(-1, 1) + weights["bs"])
        yn = np.asarray(yn).ravel()
    return atual + yn * weights["au"] + weights["bu"]


@dataclass
class SplitMetrics:
    n: int
    nash: float
    pers: float
    mae: float
    e95: float
    rmse: float


def compute_metrics(y_true: np.ndarray, y_pred: np.ndarray, y_pers: np.ndarray) -> SplitMetrics:
    err = y_pred - y_true
    mse = float(np.mean(err**2))
    mse_p = float(np.mean((y_pers - y_true) ** 2))
    den = float(np.sum((y_true - np.mean(y_true)) ** 2))
    nash = float(1.0 - np.sum(err**2) / den) if den > 0 else float("nan")
    pers = float(1.0 - mse / mse_p) if mse_p > 0 else float("nan")
    abs_err = np.abs(err)
    return SplitMetrics(
        n=int(y_true.size),
        nash=nash,
        pers=pers,
        mae=float(np.mean(abs_err)),
        e95=float(np.percentile(abs_err, 95)),
        rmse=float(np.sqrt(mse)),
    )


class MimoMLP:
    """MLP multi-saída estilo PREVINE: logsig na oculta e em cada saída."""

    def __init__(self, n_inputs: int, n_outputs: int, n_hidden: int, seed: int = 42):
        rng = np.random.default_rng(seed)
        self.n_inputs = n_inputs
        self.n_outputs = n_outputs
        self.n_hidden = n_hidden
        self.wh = rng.normal(0, 0.5, size=(n_hidden, n_inputs))
        self.bh = np.zeros(n_hidden)
        self.ws = rng.normal(0, 0.5, size=(n_outputs, n_hidden))
        self.bs = np.zeros(n_outputs)
        self.x_mean = np.zeros(n_inputs)
        self.x_std = np.ones(n_inputs)
        self.y_mean = np.zeros(n_outputs)
        self.y_std = np.ones(n_outputs)

    def _norm_x(self, x: np.ndarray) -> np.ndarray:
        return (x - self.x_mean) / self.x_std

    def _norm_y(self, y: np.ndarray) -> np.ndarray:
        return (y - self.y_mean) / self.y_std

    def forward_delta(self, x: np.ndarray) -> np.ndarray:
        x = np.atleast_2d(x)
        pn = self._norm_x(x)
        h = logsig(pn @ self.wh.T + self.bh)
        yn = logsig(h @ self.ws.T + self.bs)
        return yn * self.y_std + self.y_mean

    def fit(
        self,
        x_train: np.ndarray,
        y_train: np.ndarray,
        x_val: np.ndarray,
        y_val: np.ndarray,
        *,
        max_epochs: int = 400,
        batch_size: int = 64,
        lr: float = 0.01,
        patience: int = 30,
        seed: int = 42,
        horizon_weights: np.ndarray | None = None,
        trajectory_weight: float = 0.0,
        rising_mono_weight: float = 0.0,
    ) -> dict:
        rng = np.random.default_rng(seed)
        self.x_mean = x_train.mean(axis=0)
        self.x_std = np.clip(x_train.std(axis=0), 1e-6, None)
        self.y_mean = y_train.mean(axis=0)
        self.y_std = np.clip(y_train.std(axis=0), 1e-6, None)
        if horizon_weights is None:
            horizon_weights = np.ones(self.n_outputs, dtype=float)
        else:
            horizon_weights = np.asarray(horizon_weights, float)

        best_state = None
        best_val = float("inf")
        stale = 0
        history: list[float] = []

        n = x_train.shape[0]
        for epoch in range(max_epochs):
            order = rng.permutation(n)
            for start in range(0, n, batch_size):
                idx = order[start : start + batch_size]
                xb = x_train[idx]
                yb = y_train[idx]
                self._train_batch(
                    xb,
                    yb,
                    lr,
                    horizon_weights=horizon_weights,
                    trajectory_weight=trajectory_weight,
                    rising_mono_weight=rising_mono_weight,
                )

            val_pred = self.forward_delta(x_val)
            val_loss = float(np.mean(horizon_weights * (val_pred - y_val) ** 2))
            if trajectory_weight > 0 and self.n_outputs >= 2:
                true_shape = y_val[:, 1] - y_val[:, 0]
                pred_shape = val_pred[:, 1] - val_pred[:, 0]
                val_loss += trajectory_weight * float(np.mean((pred_shape - true_shape) ** 2))
            history.append(val_loss)
            if val_loss + 1e-6 < best_val:
                best_val = val_loss
                stale = 0
                best_state = (
                    self.wh.copy(),
                    self.bh.copy(),
                    self.ws.copy(),
                    self.bs.copy(),
                    self.x_mean.copy(),
                    self.x_std.copy(),
                    self.y_mean.copy(),
                    self.y_std.copy(),
                )
            else:
                stale += 1
                if stale >= patience:
                    break

        if best_state is not None:
            (
                self.wh,
                self.bh,
                self.ws,
                self.bs,
                self.x_mean,
                self.x_std,
                self.y_mean,
                self.y_std,
            ) = best_state
        return {
            "epochs": len(history),
            "best_val_mse": best_val,
            "history_tail": history[-5:],
            "trajectory_weight": trajectory_weight,
            "rising_mono_weight": rising_mono_weight,
            "horizon_weights": horizon_weights.tolist(),
        }

    def _train_batch(
        self,
        xb: np.ndarray,
        yb: np.ndarray,
        lr: float,
        *,
        horizon_weights: np.ndarray,
        trajectory_weight: float,
        rising_mono_weight: float,
    ) -> None:
        xn = self._norm_x(xb)
        pn = xn
        zh = pn @ self.wh.T + self.bh
        h = logsig(zh)
        zo = h @ self.ws.T + self.bs
        y_hat_n = logsig(zo)
        y_hat = y_hat_n * self.y_std + self.y_mean
        err = y_hat - yb
        loss_grad_y = 2.0 * horizon_weights * err / xb.shape[0]

        if self.n_outputs >= 2 and trajectory_weight > 0:
            true_shape = yb[:, 1] - yb[:, 0]
            pred_shape = y_hat[:, 1] - y_hat[:, 0]
            shape_err = pred_shape - true_shape
            # d(loss)/d(y0) = -2 * shape_err / n; d(loss)/d(y1) = +2 * shape_err / n
            g = 2.0 * trajectory_weight * shape_err / xb.shape[0]
            loss_grad_y[:, 0] -= g
            loss_grad_y[:, 1] += g

        if self.n_outputs >= 2 and rising_mono_weight > 0:
            # When true trajectory is rising, penalize pred4 < pred2.
            rising = (yb[:, 1] - yb[:, 0]) > 0
            violation = np.maximum(0.0, y_hat[:, 0] - y_hat[:, 1])
            # d ReLU(y0-y1)/dy0 = 1 if violation>0 else 0; /dy1 = -1
            active = rising & (violation > 0)
            if np.any(active):
                g = rising_mono_weight / xb.shape[0]
                loss_grad_y[active, 0] += g
                loss_grad_y[active, 1] -= g

        dyhat_dyn = self.y_std
        dyn_dzo = y_hat_n * (1.0 - y_hat_n)
        dzo = loss_grad_y * dyhat_dyn * dyn_dzo

        dws = dzo.T @ h
        dbs = dzo.sum(axis=0)
        dh = dzo @ self.ws
        dzh = dh * h * (1.0 - h)
        dwh = dzh.T @ pn
        dbh = dzh.sum(axis=0)

        self.ws -= lr * dws
        self.bs -= lr * dbs
        self.wh -= lr * dwh
        self.bh -= lr * dbh


def rising_inconsistency_rate(y_true_abs: np.ndarray, y_pred_abs: np.ndarray) -> dict:
    """Fraction of rising cases where predicted 4h level < predicted 2h level."""
    if y_true_abs.ndim != 2 or y_true_abs.shape[1] < 2:
        return {"rate": None, "n_rising": 0, "n_violations": 0}
    rising = y_true_abs[:, 1] > y_true_abs[:, 0]
    n_rising = int(np.sum(rising))
    if n_rising == 0:
        return {"rate": None, "n_rising": 0, "n_violations": 0}
    viol = np.sum(rising & (y_pred_abs[:, 1] < y_pred_abs[:, 0]))
    return {"rate": float(viol / n_rising), "n_rising": n_rising, "n_violations": int(viol)}


def evaluate_strategy(
    *,
    name: str,
    rows: list[dict],
    datasets: list[HorizonDataset],
    input_dataset_idx: int,
    output_specs: list[tuple[int, int]],
    predict_fn,
) -> dict:
    """output_specs: list of (dataset_idx, output_index_if_multi)."""
    by_split: dict[int, dict[str, list[float]]] = {1: {}, 2: {}, 3: {}}
    for split_id in by_split:
        for ds_idx, _ in output_specs:
            by_split[split_id][datasets[ds_idx].name] = {"true": [], "pred": [], "pers": []}

    for row in rows:
        split_id = row["split"]
        indices = row["indices"]
        in_ds = datasets[input_dataset_idx]
        x = in_ds.inputs[indices[input_dataset_idx]]
        atual_in = in_ds.atual[indices[input_dataset_idx]]
        preds = predict_fn(x, atual_in, row)
        for out_i, (ds_idx, _) in enumerate(output_specs):
            ds = datasets[ds_idx]
            idx = indices[ds_idx]
            y_true = float(ds.target_abs[idx])
            y_pers = float(ds.atual[idx])
            by_split[split_id][ds.name]["true"].append(y_true)
            by_split[split_id][ds.name]["pred"].append(float(preds[out_i]))
            by_split[split_id][ds.name]["pers"].append(y_pers)

    out = {"strategy": name, "splits": {}, "horizons": {}}
    for split_id, label in [(1, "treino"), (2, "validacao"), (3, "teste")]:
        split_payload = {}
        for ds_idx, _ in output_specs:
            ds_name = datasets[ds_idx].name
            payload = by_split[split_id][ds_name]
            if not payload["true"]:
                continue
            y_true = np.asarray(payload["true"], float)
            y_pred = np.asarray(payload["pred"], float)
            y_pers = np.asarray(payload["pers"], float)
            m = compute_metrics(y_true, y_pred, y_pers)
            split_payload[ds_name] = m.__dict__
        out["splits"][label] = split_payload

    for ds_idx, _ in output_specs:
        ds_name = datasets[ds_idx].name
        all_true = []
        all_pred = []
        all_pers = []
        for split_id in (1, 2, 3):
            payload = by_split[split_id][ds_name]
            all_true.extend(payload["true"])
            all_pred.extend(payload["pred"])
            all_pers.extend(payload["pers"])
        if all_true:
            m = compute_metrics(np.asarray(all_true), np.asarray(all_pred), np.asarray(all_pers))
            out["horizons"][ds_name] = m.__dict__
    return out


def compare_delta(a: dict, b: dict, horizon: str, metric: str = "nash") -> float:
    return float(b["splits"]["teste"][horizon][metric] - a["splits"]["teste"][horizon][metric])


def save_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
