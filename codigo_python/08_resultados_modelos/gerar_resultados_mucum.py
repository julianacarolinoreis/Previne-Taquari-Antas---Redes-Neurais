"""
Gera a estrutura de resultados de Mucum (graficos por evento + downloads de
.mat/planilha auditavel), no mesmo padrao ja usado para Santa Tereza.

Fontes de dados (nunca assumidas — sempre lidas do disco a cada execucao):
  - Planilhao consolidado: redes_neurais/mucum/PLANILHAO_TECNICO_MUCUM_CONSOLIDADO_PADRAO_STZ_ATUALIZADO.xlsx
    (aba TODOS_MODELOS) — fonte principal, cobre todas as rodadas ja processadas.
  - Rodadas mais novas que ainda nao passaram pelo atualizador do planilhao
    (ex.: uma rodada em andamento) sao lidas direto de
    status/fila_mucum.csv + resultados_mucum_rodadaN.csv de cada pasta
    RNA_MUC_MELHORIA_RODADA*, para o robo nao ficar "cego" a rodadas novas.

Criterio de selecao: score_equilibrio = min(PERS_treino, PERS_validacao, PERS_teste) > 0.75,
com 1 modelo por (horizonte, tipo, combo_id) — o de maior equilibrio.
Se um horizonte nao tiver nenhum modelo acima do corte, ele fica de fora
(nao ha fallback silencioso para um modelo pior).

Saida:
  - assets/mat/<modelo>.mat                      (copiado)
  - assets/audit_workbooks/<FAMILIA>__<modelo>.xlsx  (copiado)
  - assets/data/mucum_auditaveis_series.json     (series por evento, para os graficos)
  - index.html: <script id="data-mucum"> ganha mat_url/wb_url/mat{} nos
    modelos selecionados (patch pontual, resto do payload preservado).

Idempotente: pode rodar de novo a qualquer momento (ex.: apos cada rodada de
treino terminar) sem duplicar trabalho — arquivos ja copiados com o mesmo
tamanho sao pulados.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import scipy.io as sio

MUCUM_ROOT = Path(r"D:\PREVINE\redes_neurais\mucum")
PLANILHAO = MUCUM_ROOT / "PLANILHAO_TECNICO_MUCUM_CONSOLIDADO_PADRAO_STZ_ATUALIZADO.xlsx"
REPO = Path(r"D:\PREVINE\repo_site")
INDEX_HTML = REPO / "index.html"
MAT_DIR = REPO / "assets" / "mat"
WORKBOOK_DIR = REPO / "assets" / "audit_workbooks"
SERIES_JSON = REPO / "assets" / "data" / "mucum_auditaveis_series.json"

RAW_BASE = "https://raw.githubusercontent.com/julianacarolinoreis/Previne-Taquari-Antas---Redes-Neurais"
EQUILIBRIO_MIN = 0.75
SET_CODE = {"Treino": 0, "Validacao": 1, "Teste": 2, "Outro": 3}


def calc_score(values):
    vals = [v for v in values if v is not None]
    if not vals:
        return None, None, None
    return min(vals), sum(vals) / len(vals), max(vals) - min(vals)


def as_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_planilhao_rows():
    wb = openpyxl.load_workbook(PLANILHAO, read_only=True, data_only=True)
    ws = wb["TODOS_MODELOS"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    return [dict(zip(header, r)) for r in rows[1:]], {r[header.index("fonte_rodada")] for r in rows[1:] if r[header.index("fonte_rodada")]}


def read_csv_rows(path, delimiter_candidates=(";", ",")):
    raw = path.read_text(encoding="utf-8-sig")
    delim = ";" if raw.count(";") >= raw.count(",") else ","
    return list(csv.DictReader(raw.splitlines(), delimiter=delim))


def scan_new_rodadas(known_fontes):
    """Cobre rodadas RNA_MUC_MELHORIA_RODADA* que o planilhao ainda nao processou."""
    extra_rows = []
    for run_dir in sorted(MUCUM_ROOT.glob("RNA_MUC_MELHORIA_RODADA*")):
        if run_dir.name in known_fontes or not run_dir.is_dir():
            continue
        fila = run_dir / "status" / "fila_mucum.csv"
        n = re.search(r"RODADA(\d+)", run_dir.name)
        resultados = run_dir / f"resultados_mucum_rodada{n.group(1)}.csv" if n else None
        if not fila.exists() or not resultados or not resultados.exists():
            continue
        queue = {q["run_id"]: q for q in read_csv_rows(fila)}
        results = {r["RUN_ID"]: r for r in read_csv_rows(resultados)}
        for run_id, r in results.items():
            q = queue.get(run_id, {})
            pers_treino = as_float(r.get("PERS_TREINO"))
            pers_val = as_float(r.get("PERS_VALIDACAO"))
            pers_teste = as_float(r.get("PERS_TESTE"))
            score, media, spread = calc_score([pers_treino, pers_val, pers_teste])
            horizonte_h = q.get("horizonte_h") or ""
            tipo = (q.get("tipo_modelo") or r.get("TIPO") or "").lower()
            combo_id = q.get("modelo_id") or ""
            auditavel = run_dir / "auditaveis" / f"{run_id}_auditavel.xlsx"
            extra_rows.append({
                "modelo": run_id,
                "combo_id": combo_id,
                "horizonte": f"{horizonte_h}h" if horizonte_h else "",
                "tipo": tipo,
                "familia": f"{horizonte_h}H_{tipo.upper()}" if horizonte_h and tipo else "",
                "fonte_rodada": run_dir.name,
                "status_modelo": "CONCLUIDO",
                "score_equilibrio": score,
                "media_pers": media,
                "spread_pers": spread,
                "PERS_geral": as_float(r.get("PERS_GERAL")),
                "PERS_treino": pers_treino,
                "PERS_validacao": pers_val,
                "PERS_teste": pers_teste,
                "MAE_teste_cm": as_float(r.get("MAE_TESTE")),
                "E95_teste_cm": as_float(r.get("E95_TESTE")),
                "arquivo_mat": r.get("arquivo_mat") or str(run_dir / "mat" / f"{run_id}.mat"),
                "arquivo_auditavel": str(auditavel),
            })
    return extra_rows


def selecionar_qualificados():
    rows, known_fontes = read_planilhao_rows()
    rows += scan_new_rodadas(known_fontes)

    qualified = [
        r for r in rows
        if r.get("score_equilibrio") is not None
        and r["score_equilibrio"] > EQUILIBRIO_MIN
        and r.get("status_modelo") in ("OK", "CONCLUIDO")
    ]

    groups = {}
    for r in qualified:
        key = (r.get("horizonte"), r.get("tipo"), r.get("combo_id"))
        groups.setdefault(key, []).append(r)

    best = []
    for items in groups.values():
        items.sort(key=lambda r: (r["score_equilibrio"], str(r.get("fonte_rodada") or "")), reverse=True)
        best.append(items[0])
    best.sort(key=lambda r: (r.get("horizonte") or "", -(r["score_equilibrio"])))
    return best


def logsig(x):
    return 1.0 / (1.0 + np.exp(-x))


def validar_mat(mat_path):
    """Reconstroi pred_target_tot a partir dos pesos e sinaliza incompatibilidade em vez de assumir."""
    try:
        m = sio.loadmat(mat_path, squeeze_me=True, struct_as_record=False)
    except Exception as e:  # noqa: BLE001
        return False, f"nao foi possivel abrir o .mat: {e}", None
    required = ["wh", "bh", "ws", "bs", "ae", "be", "au", "bu", "ptot", "pred_target_tot"]
    missing = [k for k in required if k not in m]
    if missing:
        return False, f"campos ausentes: {missing}", None
    hidden = logsig(m["wh"] @ m["ptot"] + m["bh"][:, None])
    out_norm = logsig(m["ws"] @ hidden + m["bs"])
    out_real = out_norm * m["au"] + m["bu"]
    rmse = float(np.sqrt(np.mean((out_real - m["pred_target_tot"]) ** 2)))
    if rmse > 1e-6:
        return False, f"RMSE de reconstrucao alto ({rmse:.6f}) — pesos nao batem com pred_target_tot", m
    return True, "ok", m


def extrair_mat_meta(m, mat_path):
    return {
        "nh": int(m["nh"]),
        "nit": int(m["nit"]),
        "cic": int(m["Cic"]),
        "j": float(m["J"]),
        "NASH": float(m["nash_te"]),
        "NASH_VAL": float(m["nash_v"]),
        "PERS": float(m["pers_te"]),
        "e95": float(m["e95_te"]),
        "emed": float(m["mae_te"]),
        "wh": list(m["wh"].shape),
        "bh": list(m["bh"].shape),
        "size": mat_path.stat().st_size,
        "mod": datetime.fromtimestamp(mat_path.stat().st_mtime).isoformat(timespec="seconds"),
    }


def copiar_se_necessario(origem: Path, destino: Path):
    if destino.exists() and destino.stat().st_size == origem.stat().st_size:
        return False
    destino.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(origem, destino)
    return True


def montar_series_do_auditavel(auditavel_path: Path):
    wb = openpyxl.load_workbook(auditavel_path, read_only=True, data_only=True)
    ws = wb["DADOS"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {name: i for i, name in enumerate(header)}
    series = {}
    for row in rows[1:]:
        evento = row[idx["EVENTO"]]
        conjunto = row[idx["CONJUNTO"]]
        if evento is None or conjunto is None:
            continue
        dt = row[idx["DATA_HORA"]]
        ts = dt.strftime("%Y-%m-%d %H:%M") if hasattr(dt, "strftime") else str(dt)
        obs = row[idx["OBSERVADO_CM"]]
        rna = row[idx["RNA_CM"]]
        pers = row[idx["PERS_CM"]]
        erro = row[idx["ERRO_RNA_CM"]]
        erro_abs = row[idx["ERRO_ABS_CM"]]
        key = f"{evento}|{conjunto}"
        series.setdefault(key, []).append([
            ts, obs, rna, SET_CODE.get(conjunto, 3), pers, erro, erro_abs,
        ])
    return series


def ler_inputs_do_auditavel(auditavel_path: Path):
    wb = openpyxl.load_workbook(auditavel_path, read_only=True, data_only=True)
    ws = wb["INPUTS"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx_ordem = header.index("ordem")
    idx_input = header.index("input")
    pares = [(row[idx_ordem], row[idx_input]) for row in rows[1:] if row[idx_input]]
    pares.sort(key=lambda p: p[0])
    return [nome for _, nome in pares]


def montar_metrics(r):
    return {k: r.get(k) for k in (
        "PERS_geral", "PERS_treino", "PERS_validacao", "PERS_teste",
        "MAE_geral_cm", "MAE_validacao_cm", "MAE_teste_cm",
        "E95_geral_cm", "E95_validacao_cm", "E95_teste_cm",
        "NASH_validacao_csv", "NASH_teste_csv", "n_inputs", "neuronios",
    ) if r.get(k) is not None}


def gerar():
    qualificados = selecionar_qualificados()
    print(f"Modelos qualificados (score_equilibrio > {EQUILIBRIO_MIN}, deduplicados por combo/horizonte): {len(qualificados)}")

    series_json_models = []
    payload_updates = {}
    avisos = []

    for r in qualificados:
        modelo = r["modelo"]
        mat_origem = Path(r["arquivo_mat"])
        auditavel_origem = Path(r["arquivo_auditavel"])
        familia = (r.get("familia") or f"{r['horizonte'].upper()}_{r['tipo'].upper()}")

        if not mat_origem.exists():
            avisos.append(f"{modelo}: .mat nao encontrado em {mat_origem} — pulado")
            continue
        if not auditavel_origem.exists():
            avisos.append(f"{modelo}: planilha auditavel nao encontrada em {auditavel_origem} — pulado")
            continue

        ok, motivo, m = validar_mat(mat_origem)
        if not ok:
            avisos.append(f"{modelo}: {motivo} — pulado (nao publicado sem validacao)")
            continue

        mat_destino = MAT_DIR / f"{modelo}.mat"
        wb_destino = WORKBOOK_DIR / f"{familia}__{modelo}.xlsx"
        copiado_mat = copiar_se_necessario(mat_origem, mat_destino)
        copiado_wb = copiar_se_necessario(auditavel_origem, wb_destino)

        mat_meta = extrair_mat_meta(m, mat_destino)
        series = montar_series_do_auditavel(auditavel_origem)

        mat_url = f"assets/mat/{modelo}.mat"
        wb_url = f"assets/audit_workbooks/{familia}__{modelo}.xlsx"

        series_json_models.append({
            "id": modelo,
            "name": modelo,
            "family": familia,
            "horizon": r["horizonte"],
            "type": r["tipo"],
            "combo_id": r.get("combo_id"),
            "rotation": r.get("rotacao"),
            "fonte_rodada": r.get("fonte_rodada"),
            "matFile": mat_destino.name,
            "matSourceRef": str(mat_origem),
            "matUrl": mat_url,
            "matSize": mat_meta["size"],
            "workbookFile": wb_destino.name,
            "workbookUrl": wb_url,
            "sourceRef": str(auditavel_origem),
            "metrics": montar_metrics(r),
            "series": series,
        })

        payload_updates[modelo] = {
            "mat_url": mat_url, "wb_url": wb_url, "mat": mat_meta,
            "row": r, "auditavel_origem": auditavel_origem,
        }

        status_bits = []
        if copiado_mat:
            status_bits.append("mat copiado")
        if copiado_wb:
            status_bits.append("xlsx copiado")
        print(f"  OK  {modelo:40s} eq={r['score_equilibrio']:.4f}  {' + '.join(status_bits) or '(ja estava tudo copiado)'}")

    for a in avisos:
        print(f"  AVISO  {a}")

    SERIES_JSON.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "meta": {
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "modelCount": len(series_json_models),
            "criterio_selecao": f"score_equilibrio (min PERS treino/validacao/teste) > {EQUILIBRIO_MIN}, 1 por combo/horizonte",
            "setLabels": ["Treino", "Validacao", "Teste", "Outro"],
        },
        "models": series_json_models,
        "skipped": avisos,
    }
    SERIES_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nEscrito {SERIES_JSON} com {len(series_json_models)} modelos.")

    patch_index_html(payload_updates)
    return len(series_json_models), len(avisos)


ENTRY_FIELDS = (
    "familia", "horizonte", "tipo", "rotacao", "modelo", "combo_id",
    "evento_teste", "eventos_validacao", "fonte_rodada", "status_modelo",
    "n_inputs", "neuronios", "nit", "ciclos", "N_geral", "N_treino",
    "N_validacao", "N_teste", "PERS_geral", "PERS_treino", "PERS_validacao",
    "PERS_teste", "score_equilibrio", "MAE_geral_cm", "MAE_validacao_cm",
    "MAE_teste_cm", "E95_geral_cm", "E95_validacao_cm", "E95_teste_cm",
    "NASH_validacao_csv", "NASH_teste_csv", "correlacao_teste_csv",
    "arquivo_auditavel", "arquivo_mat", "J",
)


def montar_inputs_index(nomes_input, inputs_list):
    idx = []
    for nome in nomes_input:
        if nome not in inputs_list:
            inputs_list.append(nome)
        idx.append(inputs_list.index(nome))
    return idx


def patch_index_html(payload_updates: dict):
    if not payload_updates:
        return
    html = INDEX_HTML.read_text(encoding="utf-8")
    m = re.search(r'(<script id="data-mucum"[^>]*>)(.*?)(</script>)', html, re.S)
    if not m:
        print("AVISO: <script id=\"data-mucum\"> nao encontrado em index.html — payload nao foi atualizado.")
        return
    data = json.loads(m.group(2))
    by_modelo = {model.get("modelo"): model for model in data.get("models", [])}

    atualizados = 0
    criados = 0
    for modelo, upd in payload_updates.items():
        model = by_modelo.get(modelo)
        if model is not None:
            model["mat_url"] = upd["mat_url"]
            model["wb_url"] = upd["wb_url"]
            model["mat"] = upd["mat"]
            atualizados += 1
            continue

        r = upd["row"]
        nomes_input = ler_inputs_do_auditavel(upd["auditavel_origem"])
        novo = {k: r.get(k) for k in ENTRY_FIELDS}
        novo["modelo"] = modelo
        novo["inputs"] = montar_inputs_index(nomes_input, data["inputs"])
        novo["mat_url"] = upd["mat_url"]
        novo["wb_url"] = upd["wb_url"]
        novo["mat"] = upd["mat"]
        data["models"].append(novo)
        criados += 1

    novo_json = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    novo_html = html[:m.start(2)] + novo_json + html[m.end(2):]
    INDEX_HTML.write_text(novo_html, encoding="utf-8")
    print(f"index.html: {atualizados} modelo(s) existentes atualizados + {criados} entrada(s) nova(s) criadas no payload data-mucum.")


def pin_urls_para_commit(sha: str):
    """Troca URLs relativas (assets/mat/..., assets/audit_workbooks/...) por
    raw.githubusercontent.com fixado no commit informado — necessario porque
    o index.html resolve toda URL nao-absoluta contra um commit antigo
    pinado (WB_RAW), entao um arquivo novo so fica baixavel com uma URL
    absoluta apontando para o commit que de fato o contem."""

    def pin(url):
        if not url or url.startswith("http"):
            return url
        return f"{RAW_BASE}/{sha}/{url}"

    data = json.loads(SERIES_JSON.read_text(encoding="utf-8"))
    for model in data["models"]:
        model["matUrl"] = pin(model.get("matUrl"))
        model["workbookUrl"] = pin(model.get("workbookUrl"))
    SERIES_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    html = INDEX_HTML.read_text(encoding="utf-8")
    m = re.search(r'(<script id="data-mucum"[^>]*>)(.*?)(</script>)', html, re.S)
    payload = json.loads(m.group(2))
    n = 0
    for model in payload["models"]:
        if (model.get("mat_url") or "").startswith("assets/") or (model.get("wb_url") or "").startswith("assets/"):
            model["mat_url"] = pin(model.get("mat_url"))
            model["wb_url"] = pin(model.get("wb_url"))
            n += 1
    novo_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    novo_html = html[:m.start(2)] + novo_json + html[m.end(2):]
    INDEX_HTML.write_text(novo_html, encoding="utf-8")
    print(f"URLs fixadas no commit {sha}: {len(data['models'])} em mucum_auditaveis_series.json, {n} em index.html.")


if __name__ == "__main__":
    if len(sys.argv) > 2 and sys.argv[1] == "--pin-commit":
        pin_urls_para_commit(sys.argv[2])
        sys.exit(0)
    n_ok, n_avisos = gerar()
    if n_avisos:
        print(f"\n{n_avisos} aviso(s) acima — modelos nao publicados sem validacao passar.")
    sys.exit(0)
